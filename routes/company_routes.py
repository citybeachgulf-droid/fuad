"""Blueprint for valuation company portal routes and templates."""
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, jsonify
from sqlalchemy import or_
from flask_login import login_required, current_user
from models import db, ValuationRequest, CompanyProfile, CompanyContact, VisitAppointment, Conversation, Message, ActivityLog, CompanyLandPrice
from werkzeug.utils import secure_filename
from utils import store_file_and_get_url
import os
import time

company_bp = Blueprint('company', __name__, template_folder='templates/company')

@company_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    # جلب الطلبات المرتبطة بالشركة
    # اعرض فقط المعاملات الجديدة (pending)
    requests = (
        ValuationRequest.query
        .filter(
            ValuationRequest.company_id == current_user.id,
            (ValuationRequest.status == 'pending')
        )
        .order_by(ValuationRequest.id.desc())
        .all()
    )
    # إحصاءات سريعة لحالات معاملات الشركة
    counts_rows = (
        db.session.query(ValuationRequest.status, db.func.count(ValuationRequest.id))
        .filter(ValuationRequest.company_id == current_user.id)
        .group_by(ValuationRequest.status)
        .all()
    )
    counts_map = {status or 'unknown': cnt for status, cnt in counts_rows}
    status_counts = {
        'pending': counts_map.get('pending', 0),
        'revision_requested': counts_map.get('revision_requested', 0),
        'rejected': counts_map.get('rejected', 0),
        'completed': counts_map.get('completed', 0),
    }
    # المواعيد المقترحة من العميل والتي ما زالت بانتظار موافقة الشركة
    pending_appts = (
        VisitAppointment.query
        .join(ValuationRequest, VisitAppointment.valuation_request_id == ValuationRequest.id)
        .filter(
            ValuationRequest.company_id == current_user.id,
            VisitAppointment.status == 'pending',
            VisitAppointment.proposed_by == 'client',
        )
        .order_by(VisitAppointment.created_at.desc())
        .all()
    )
    profile = CompanyProfile.query.filter_by(user_id=current_user.id).first()
    return render_template(
        'company/dashboard.html',
        requests=requests,
        profile=profile,
        pending_appointments=pending_appts,
        status_counts=status_counts,
    )


# صفحة تعرض معاملات الشركة حسب الحالة (مرفوضة / مستندات ناقصة / تم التثمين)
@company_bp.route('/transactions/status')
@login_required
def transactions_status():
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403

    # تبويب نشط من الاستعلام
    active_tab = (request.args.get('tab') or 'rejected').lower()
    if active_tab not in {'rejected', 'missing_docs', 'completed'}:
        active_tab = 'rejected'

    base = ValuationRequest.query.filter(ValuationRequest.company_id == current_user.id)

    # القوائم حسب الحالة
    rejected_list = base.filter(ValuationRequest.status == 'rejected').order_by(ValuationRequest.id.desc()).all()
    missing_docs_list = base.filter(ValuationRequest.status == 'revision_requested').order_by(ValuationRequest.id.desc()).all()
    completed_list = base.filter(ValuationRequest.status == 'completed').order_by(ValuationRequest.id.desc()).all()

    counts = {
        'rejected': len(rejected_list),
        'missing_docs': len(missing_docs_list),
        'completed': len(completed_list),
    }

    # اختر القائمة حسب التبويب
    if active_tab == 'rejected':
        active_list = rejected_list
    elif active_tab == 'missing_docs':
        active_list = missing_docs_list
    else:
        active_list = completed_list

    return render_template(
        'company/transactions_status.html',
        active_tab=active_tab,
        items=active_list,
        counts=counts,
    )

@company_bp.route('/submit/<int:request_id>', methods=['GET', 'POST'])
@login_required
def submit_valuation(request_id):
    vr = ValuationRequest.query.get(request_id)
    if request.method == 'POST':
        vr.value = float(request.form.get('value') or 0)
        vr.status = 'completed'
        db.session.commit()
        flash('Valuation submitted', 'success')
        return redirect(url_for('company.dashboard'))
    return render_template('company/submit.html', request_obj=vr)


@company_bp.route('/requests/<int:request_id>')
@login_required
def request_detail(request_id: int):
    """عرض تفاصيل معاملة التثمين للشركة الحالية."""
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    req = ValuationRequest.query.get_or_404(request_id)
    if req.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403
    # احضار المواعيد المرتبطة بأحدث ترتيب
    appts = VisitAppointment.query.filter_by(valuation_request_id=req.id).order_by(VisitAppointment.created_at.desc()).all()
    return render_template('company/request_detail.html', request_obj=req, appointments=appts)


@company_bp.route('/requests/<int:request_id>/upload_report', methods=['POST'])
@login_required
def upload_final_report(request_id: int):
    """Upload final report (PDF/image) after client acceptance and send to client via conversation."""
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    req = ValuationRequest.query.get_or_404(request_id)
    if req.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    # Only allow after client has accepted the valuation
    if (req.status or '').lower() not in { 'approved' }:
        flash('يمكن رفع التقرير بعد موافقة العميل على التقييم.', 'warning')
        return redirect(url_for('company.request_detail', request_id=req.id))

    file = request.files.get('final_report')
    if not file or not file.filename:
        flash('يرجى اختيار ملف التقرير النهائي', 'danger')
        return redirect(url_for('company.request_detail', request_id=req.id) + '#final-report')

    # Accept only PDF/images similar to client docs
    allowed_exts = { 'pdf', 'png', 'jpg', 'jpeg', 'webp', 'gif' }
    ext = (os.path.splitext(file.filename)[1].lower().lstrip('.') or '')
    if ext not in allowed_exts:
        flash('نوع الملف غير مدعوم. المسموح: PDF أو صور', 'danger')
        return redirect(url_for('company.request_detail', request_id=req.id) + '#final-report')

    # Store under static/uploads/reports/req_<id>/
    static_root = os.path.join(current_app.root_path, 'static')
    reports_root = os.path.join(static_root, 'uploads', 'reports', f'req_{req.id}')
    os.makedirs(reports_root, exist_ok=True)

    safe_name = secure_filename(file.filename)
    ts = int(time.time())
    filename = f"final_report_{ts}_{safe_name}"
    object_key = f"uploads/reports/req_{req.id}/{filename}"
    stored = store_file_and_get_url(
        file,
        key=object_key,
        local_abs_dir=reports_root,
        filename=filename,
    )

    if not stored:
        flash('تعذّر حفظ التقرير. حاول مرة أخرى.', 'danger')
        return redirect(url_for('company.request_detail', request_id=req.id) + '#final-report')

    # Persist as RequestDocument of type 'final_report'
    from models import RequestDocument
    doc = RequestDocument(
        valuation_request_id=req.id,
        doc_type='final_report',
        file_path=stored,
        original_filename=safe_name,
    )
    db.session.add(doc)

    # Notify client via conversation
    conv = Conversation.query.filter_by(client_id=req.client_id, company_id=current_user.id).first()
    if not conv:
        conv = Conversation(client_id=req.client_id, company_id=current_user.id, status='open')
        db.session.add(conv)
        db.session.flush()
        db.session.add(ActivityLog(conversation_id=conv.id, actor_id=current_user.id, action='conversation_created'))

    report_url = stored if (stored.lower().startswith('http://') or stored.lower().startswith('https://')) else url_for('static', filename=stored, _external=True)
    msg_content = f"تم رفع التقرير النهائي لطلب التثمين #{req.id}.\nيمكنك الاطلاع على التقرير هنا: {report_url}"
    db.session.add(Message(conversation_id=conv.id, sender_id=current_user.id, content=msg_content))
    db.session.add(ActivityLog(conversation_id=conv.id, actor_id=current_user.id, action='message_sent'))

    db.session.commit()
    flash('تم رفع التقرير النهائي وإرساله للعميل', 'success')
    return redirect(url_for('company.request_detail', request_id=req.id) + '#final-report')


@company_bp.route('/requests/<int:request_id>/reject', methods=['POST'])
@login_required
def reject_request(request_id: int):
    """رفض معاملة مع كتابة سبب يظهر للعميل والإدارة."""
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    req = ValuationRequest.query.get_or_404(request_id)
    if req.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    reason = (request.form.get('reason') or '').strip()
    if not reason:
        flash('يرجى كتابة سبب الرفض', 'danger')
        return redirect(url_for('company.request_detail', request_id=req.id))

    from datetime import datetime
    req.status = 'rejected'
    req.rejection_reason = reason
    req.rejected_at = datetime.utcnow()
    # After rejection, remove appointments and unassign from this company
    VisitAppointment.query.filter_by(valuation_request_id=req.id).delete()
    req.company_id = None

    # إرسال رسالة إلى العميل عبر نظام المحادثة
    conv = Conversation.query.filter_by(client_id=req.client_id, company_id=current_user.id).first()
    if not conv:
        conv = Conversation(client_id=req.client_id, company_id=current_user.id, status='open')
        db.session.add(conv)
        db.session.flush()
        db.session.add(ActivityLog(conversation_id=conv.id, actor_id=current_user.id, action='conversation_created'))

    content = f"تم رفض طلب التثمين #{req.id}. السبب:\n{reason}"
    db.session.add(Message(conversation_id=conv.id, sender_id=current_user.id, content=content))
    db.session.add(ActivityLog(conversation_id=conv.id, actor_id=current_user.id, action='message_sent'))

    db.session.commit()
    flash('تم رفض المعاملة مع توضيح السبب للعميل', 'success')
    return redirect(url_for('company.dashboard'))


@company_bp.route('/requests/<int:request_id>/missing-docs', methods=['POST'])
@login_required
def mark_missing_documents(request_id: int):
    """تحديد أن الطلب لديه مستندات ناقصة مع إرسال ملاحظات للعميل."""
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    req = ValuationRequest.query.get_or_404(request_id)
    if req.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    notes = (request.form.get('notes') or '').strip()
    if not notes:
        flash('يرجى كتابة الملاحظات حول المستندات الناقصة', 'danger')
        return redirect(url_for('company.request_detail', request_id=req.id) + '#missing-docs')

    # تحديث حالة الطلب
    req.status = 'revision_requested'

    # إرسال رسالة عبر نظام المحادثات إلى العميل
    conv = Conversation.query.filter_by(client_id=req.client_id, company_id=current_user.id).first()
    if not conv:
        conv = Conversation(client_id=req.client_id, company_id=current_user.id, status='open')
        db.session.add(conv)
        db.session.flush()
        db.session.add(ActivityLog(conversation_id=conv.id, actor_id=current_user.id, action='conversation_created'))

    content = f"طلب مستندات ناقصة بخصوص طلب التثمين #{req.id}:\n{notes}"
    db.session.add(Message(conversation_id=conv.id, sender_id=current_user.id, content=content))
    db.session.add(ActivityLog(conversation_id=conv.id, actor_id=current_user.id, action='message_sent'))

    db.session.commit()
    flash('تم إرسال طلب المستندات الناقصة إلى العميل مع الملاحظات', 'success')
    return redirect(url_for('company.request_detail', request_id=req.id))


@company_bp.route('/appointments/<int:appointment_id>/accept', methods=['POST'])
@login_required
def accept_appointment(appointment_id: int):
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    appt = VisitAppointment.query.get_or_404(appointment_id)
    vr = ValuationRequest.query.get_or_404(appt.valuation_request_id)
    if vr.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    appt.status = 'accepted'
    db.session.commit()
    flash('تمت الموافقة على الموعد. يمكنك تأكيده كموعد نهائي لاحقاً.', 'success')
    return redirect(url_for('company.request_detail', request_id=vr.id))


@company_bp.route('/requests/<int:request_id>/remove', methods=['POST'])
@login_required
def remove_request(request_id: int):
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    req = ValuationRequest.query.get_or_404(request_id)
    if req.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    # Unassign the request from this company and clear related appointments
    VisitAppointment.query.filter_by(valuation_request_id=req.id).delete()
    req.company_id = None
    try:
        db.session.commit()
        flash('تم حذف المعاملة من حساب الشركة', 'info')
    except Exception:
        db.session.rollback()
        flash('تعذّر حذف المعاملة من حساب الشركة', 'danger')

    return redirect(url_for('company.transactions_status', tab='rejected'))

@company_bp.route('/appointments/<int:appointment_id>/reject', methods=['POST'])
@login_required
def reject_appointment(appointment_id: int):
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    appt = VisitAppointment.query.get_or_404(appointment_id)
    vr = ValuationRequest.query.get_or_404(appt.valuation_request_id)
    if vr.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    appt.status = 'rejected'
    db.session.commit()
    flash('تم رفض الموعد. يرجى اقتراح موعد بديل.', 'info')
    return redirect(url_for('company.request_detail', request_id=vr.id))


@company_bp.route('/appointments/<int:appointment_id>/finalize', methods=['POST'])
@login_required
def finalize_appointment(appointment_id: int):
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    appt = VisitAppointment.query.get_or_404(appointment_id)
    vr = ValuationRequest.query.get_or_404(appt.valuation_request_id)
    if vr.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    # ضع جميع المواعيد الأخرى كـ rejected إذا لم تكن نهائية
    others = VisitAppointment.query.filter(
        VisitAppointment.valuation_request_id == vr.id,
        VisitAppointment.id != appt.id,
    ).all()
    for o in others:
        if o.status != 'final':
            o.status = 'rejected'

    appt.status = 'final'
    db.session.commit()
    flash('تم تحديد موعد الزيارة النهائي', 'success')
    return redirect(url_for('company.request_detail', request_id=vr.id))


@company_bp.route('/appointments/propose/<int:request_id>', methods=['POST'])
@login_required
def propose_company_appointment(request_id: int):
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403
    vr = ValuationRequest.query.get_or_404(request_id)
    if vr.company_id != current_user.id:
        return "غير مصرح لك بالوصول", 403

    from datetime import datetime
    proposed_raw = (request.form.get('proposed_time') or '').strip()
    notes = (request.form.get('notes') or '').strip() or None
    if not proposed_raw:
        flash('يرجى تحديد وقت الموعد', 'danger')
        return redirect(url_for('company.request_detail', request_id=vr.id))
    try:
        proposed_dt = datetime.fromisoformat(proposed_raw)
    except Exception:
        flash('تنسيق وقت غير صالح', 'danger')
        return redirect(url_for('company.request_detail', request_id=vr.id))

    appt = VisitAppointment(
        valuation_request_id=vr.id,
        proposed_time=proposed_dt,
        proposed_by='company',
        status='pending',
        notes=notes,
    )
    db.session.add(appt)

    # خيارياً: رفض الموعد الأصلي الذي اقترحه العميل إذا طُلب ذلك
    reject_original_id_raw = request.form.get('reject_original_id')
    if reject_original_id_raw:
        try:
            reject_id = int(reject_original_id_raw)
            original_appt = VisitAppointment.query.get(reject_id)
        except Exception:
            original_appt = None
        if original_appt and original_appt.valuation_request_id == vr.id:
            original_appt.status = 'rejected'
    db.session.commit()
    flash('تم اقتراح موعد بديل للعميل', 'success')
    return redirect(url_for('company.request_detail', request_id=vr.id))


# ================================
# إدارة أسعار الأراضي للشركة
# ================================
@company_bp.route('/land_prices', methods=['GET'])
@login_required
def land_prices():
    """عرض قائمة أسعار الأراضي التي رفعتها الشركة مع البحث والتعديل."""
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403

    profile = CompanyProfile.query.filter_by(user_id=current_user.id).first()
    if not profile:
        profile = CompanyProfile(user_id=current_user.id)
        db.session.add(profile)
        db.session.commit()

    q = (request.args.get('q') or '').strip()
    base = CompanyLandPrice.query.filter_by(company_profile_id=profile.id)
    if q:
        like = f"%{q}%"
        base = base.filter(or_(CompanyLandPrice.wilaya.ilike(like), CompanyLandPrice.region.ilike(like)))

    prices = base.order_by(CompanyLandPrice.wilaya.asc(), CompanyLandPrice.region.asc()).all()
    return render_template('company/land_prices.html', items=prices, q=q)


def _parse_float_field(raw_value):
    """Normalize Arabic/Persian digits, accept ranges, return float or None."""
    if raw_value in (None, ''):
        return None
    try:
        s = str(raw_value).strip()
        if s in {'-', '–', '—'}:
            return None
        digits_src = '٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹'
        digits_dst = '01234567890123456789'
        trans = str.maketrans({src: dst for src, dst in zip(digits_src, digits_dst)})
        s = s.translate(trans)
        s = s.replace('\u066b', '.').replace('٫', '.').replace('٬', '').replace(',', '')
        import re
        nums = re.findall(r'-?\d+(?:\.\d+)?', s)
        if not nums:
            return None
        if len(nums) == 1:
            return float(nums[0])
        a, b = float(nums[0]), float(nums[1])
        return (a + b) / 2.0
    except Exception:
        return None


@company_bp.route('/land_prices/<int:item_id>/edit', methods=['POST'])
@login_required
def edit_land_price(item_id: int):
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403

    profile = CompanyProfile.query.filter_by(user_id=current_user.id).first()
    if not profile:
        return "غير مصرح لك بالوصول", 403

    obj = CompanyLandPrice.query.get_or_404(item_id)
    if obj.company_profile_id != profile.id:
        return "غير مصرح لك بالوصول", 403

    wilaya = (request.form.get('wilaya') or obj.wilaya).strip()
    region = (request.form.get('region') or obj.region).strip()
    ph = _parse_float_field(request.form.get('price_housing'))
    pc = _parse_float_field(request.form.get('price_commercial'))
    pi = _parse_float_field(request.form.get('price_industrial'))
    pa = _parse_float_field(request.form.get('price_agricultural'))
    pps = _parse_float_field(request.form.get('price_per_sqm'))

    # التحديثات
    obj.wilaya = wilaya or obj.wilaya
    obj.region = region or obj.region
    if ph is not None:
        obj.price_housing = ph
    if pc is not None:
        obj.price_commercial = pc
    if pi is not None:
        obj.price_industrial = pi
    if pa is not None:
        obj.price_agricultural = pa
    if pps is not None:
        obj.price_per_sqm = pps

    # حافظ على price_per_sqm كقيمة افتراضية عند غيابها
    if obj.price_per_sqm is None:
        fallback = next((v for v in [obj.price_housing, obj.price_commercial, obj.price_industrial, obj.price_agricultural] if v is not None), None)
        if fallback is not None:
            obj.price_per_sqm = fallback

    try:
        db.session.commit()
        flash('تم تحديث السجل بنجاح', 'success')
    except Exception:
        db.session.rollback()
        flash('تعذّر حفظ التعديلات. تأكد من عدم تكرار الولاية/المنطقة.', 'danger')

    return redirect(url_for('company.land_prices', q=request.args.get('q') or None))


@company_bp.route('/land_prices/new', methods=['POST'])
@login_required
def create_land_price():
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403

    profile = CompanyProfile.query.filter_by(user_id=current_user.id).first()
    if not profile:
        profile = CompanyProfile(user_id=current_user.id)
        db.session.add(profile)
        db.session.commit()

    wilaya = (request.form.get('wilaya') or '').strip()
    region = (request.form.get('region') or '').strip()
    if not wilaya or not region:
        flash('يرجى إدخال الولاية والمنطقة', 'danger')
        return redirect(url_for('company.land_prices'))

    ph = _parse_float_field(request.form.get('price_housing'))
    pc = _parse_float_field(request.form.get('price_commercial'))
    pi = _parse_float_field(request.form.get('price_industrial'))
    pa = _parse_float_field(request.form.get('price_agricultural'))
    pps = _parse_float_field(request.form.get('price_per_sqm'))

    fallback = next((v for v in [ph, pc, pi, pa, pps] if v is not None), None)
    obj = CompanyLandPrice(
        company_profile_id=profile.id,
        wilaya=wilaya,
        region=region,
        price_housing=ph,
        price_commercial=pc,
        price_industrial=pi,
        price_agricultural=pa,
        price_per_sqm=pps if pps is not None else fallback,
    )

    try:
        db.session.add(obj)
        db.session.commit()
        flash('تمت إضافة السجل بنجاح', 'success')
    except Exception:
        db.session.rollback()
        flash('تعذّر إضافة السجل. تأكد من عدم تكرار الولاية/المنطقة.', 'danger')

    return redirect(url_for('company.land_prices'))


@company_bp.route('/land_prices/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_land_price(item_id: int):
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403

    profile = CompanyProfile.query.filter_by(user_id=current_user.id).first()
    if not profile:
        return redirect(url_for('company.land_prices'))

    obj = CompanyLandPrice.query.get_or_404(item_id)
    if obj.company_profile_id != profile.id:
        return "غير مصرح لك بالوصول", 403

    try:
        db.session.delete(obj)
        db.session.commit()
        flash('تم حذف السجل', 'info')
    except Exception:
        db.session.rollback()
        flash('تعذّر حذف السجل', 'danger')

    return redirect(url_for('company.land_prices'))

# ================================
# إدارة ملف تعريف الشركة
# ================================
ALLOWED_LOGO_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

def _allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_LOGO_EXTENSIONS


@company_bp.route('/profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403

    profile = CompanyProfile.query.filter_by(user_id=current_user.id).first()
    if not profile:
        profile = CompanyProfile(user_id=current_user.id)
        db.session.add(profile)
        db.session.commit()

    # جلب بيانات أرقام التواصل
    contacts_list = list(getattr(profile, 'contacts', []))

    if request.method == 'POST':
        services = request.form.get('services') or None
        about = request.form.get('about') or None
        website = request.form.get('website') or None

        # معالجة رفع الشعار
        file = request.files.get('logo')
        if file and file.filename:
            if not _allowed_file(file.filename):
                flash('صيغة الشعار غير مدعومة', 'danger')
                return render_template('company/profile_edit.html', profile=profile, contacts=contacts_list)
            upload_folder = current_app.config.get('UPLOAD_FOLDER')
            os.makedirs(upload_folder, exist_ok=True)
            filename = f"company_{current_user.id}_{int(time.time())}_" + secure_filename(file.filename)
            # B2 object key under uploads/logos
            object_key = f"uploads/logos/{filename}"
            stored = store_file_and_get_url(
                file,
                key=object_key,
                local_abs_dir=upload_folder,
                filename=filename,
            )
            profile.logo_path = stored.replace('\\', '/') if stored else None

        profile.services = services
        profile.about = about
        profile.website = website

        # تمت إزالة إدارة البنوك من ملف الشركة — البنوك هي من تعتمد الشركات

        # تحديث أرقام التواصل
        labels = request.form.getlist('contact_label')
        phones = request.form.getlist('contact_phone')
        if hasattr(profile, 'contacts'):
            profile.contacts.clear()
        for label, phone in zip(labels, phones):
            phone_value = (phone or '').strip()
            label_value = (label or '').strip() or None
            if phone_value:
                profile.contacts.append(CompanyContact(label=label_value, phone=phone_value))
        db.session.commit()
        flash('تم حفظ الملف التعريفي بنجاح', 'success')
        return redirect(url_for('company.edit_profile'))

    # تمرير البيانات إلى الواجهة
    contacts_list = list(getattr(profile, 'contacts', []))
    return render_template('company/profile_edit.html', profile=profile, contacts=contacts_list)


# ================================
# رفع أسعار الأراضي المعتمدة للشركة (Excel/CSV)
# ================================
@company_bp.route('/land_prices/upload', methods=['POST'])
@login_required
def upload_company_land_prices():
    if current_user.role != 'company':
        return "غير مصرح لك بالوصول", 403

    profile = CompanyProfile.query.filter_by(user_id=current_user.id).first()
    if not profile:
        profile = CompanyProfile(user_id=current_user.id)
        db.session.add(profile)
        db.session.commit()

    file = request.files.get('prices_file')
    if not file or not file.filename:
        flash('الرجاء اختيار ملف .xlsx أو .csv', 'danger')
        return redirect(url_for('company.edit_profile'))

    filename_lower = file.filename.lower()
    header_row = None
    rows_iter = None

    # Parse XLSX
    if filename_lower.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook  # lazy import
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                flash('ملف الإكسل فارغ.', 'danger')
                return redirect(url_for('company.edit_profile'))
        except ImportError:
            flash('مكتبة openpyxl غير متوفرة. يرجى رفع ملف CSV بدلاً من ذلك.', 'danger')
            return redirect(url_for('company.edit_profile'))
        except Exception:
            flash('تعذر قراءة ملف الإكسل. تأكد من أن الصيغة .xlsx صحيحة.', 'danger')
            return redirect(url_for('company.edit_profile'))

    # Parse CSV
    elif filename_lower.endswith('.csv'):
        import csv, io
        try:
            raw = file.stream.read()
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                try:
                    text = raw.decode('cp1256')
                except Exception:
                    text = raw.decode('latin1')
            reader = csv.reader(io.StringIO(text))
            header_row = next(reader, None)
            if header_row is None:
                flash('ملف CSV فارغ.', 'danger')
                return redirect(url_for('company.edit_profile'))
            rows_iter = reader
        except Exception:
            flash('تعذر قراءة ملف CSV.', 'danger')
            return redirect(url_for('company.edit_profile'))
    else:
        flash('صيغة غير مدعومة. الرجاء رفع .xlsx أو .csv', 'danger')
        return redirect(url_for('company.edit_profile'))

    def normalize_header_key(val):
        s = str(val or '').strip().lower()
        s = s.replace('ـ', '')
        import re
        # remove Arabic diacritics
        s = re.sub(r'[\u0617-\u061A\u064B-\u0652\u0670\u0653-\u065F]', '', s)
        # normalize common separators to spaces
        s = s.replace('_', ' ').replace('-', ' ').replace('/', ' ')
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    # Synonyms for supported headers (normalized)
    header_synonyms = {
        'wilaya': {
            'wilaya', 'الولاية', 'ولاية', 'ولايه', 'الولايه',
            'محافظة', 'المحافظة', 'governorate', 'state', 'province'
        },
        'region': {
            'region', 'المنطقة', 'منطقة', 'المنطقه', 'منطقه',
            'حي', 'الحي', 'حى', 'الحى', 'المدينة', 'مدينة',
            'district', 'area', 'neighborhood', 'neighbourhood', 'locality', 'city'
        },
        'housing': {'سكني', 'سكنية', 'سكن', 'housing'},
        'commercial': {'تجاري', 'تجارية', 'commercial'},
        'industrial': {'صناعي', 'صناعية', 'industrial'},
        'agricultural': {'زراعي', 'زراعية', 'agricultural', 'agriculture'},
        'price_per_sqm': {'price', 'price per sqm', 'price_per_sqm', 'سعر المتر', 'سعر_المتر', 'السعر'},
    }
    normalized_synonyms = {k: {normalize_header_key(v) for v in vs} for k, vs in header_synonyms.items()}

    header_map = {}
    for idx, col in enumerate(header_row):
        key = normalize_header_key(col)
        for canonical, syns in normalized_synonyms.items():
            if key in syns and canonical not in header_map:
                header_map[canonical] = idx
                break

    if 'wilaya' not in header_map or 'region' not in header_map:
        flash('العناوين يجب أن تتضمن: الولاية، المنطقة', 'danger')
        return redirect(url_for('company.edit_profile'))
    if not any(k in header_map for k in ('housing', 'commercial', 'industrial', 'agricultural', 'price_per_sqm')):
        flash('العناوين يجب أن تتضمن أحد الأعمدة: سكني، تجاري، صناعي، زراعي (أو سعر المتر القديم)', 'danger')
        return redirect(url_for('company.edit_profile'))

    inserted = 0
    updated = 0
    skipped = 0

    # تحقق من وجود عمود price_per_meter في قاعدة البيانات لملائمة الإصدارات القديمة
    try:
        from sqlalchemy import inspect as sa_inspect
        inspector = sa_inspect(db.engine)
        company_land_cols = [c['name'] for c in inspector.get_columns('company_land_prices')]
        has_price_per_meter = 'price_per_meter' in company_land_cols
    except Exception:
        has_price_per_meter = False

    for row in rows_iter:
        if row is None:
            skipped += 1
            continue

        wilaya_val = row[header_map['wilaya']] if len(row) > header_map['wilaya'] else None
        region_val = row[header_map['region']] if len(row) > header_map['region'] else None
        def get_val(col_key):
            if col_key in header_map and len(row) > header_map[col_key]:
                return row[header_map[col_key]]
            return None

        housing_val = get_val('housing')
        commercial_val = get_val('commercial')
        industrial_val = get_val('industrial')
        agricultural_val = get_val('agricultural')
        legacy_price_val = get_val('price_per_sqm')

        wilaya_str = str(wilaya_val or '').strip()
        region_str = str(region_val or '').strip()

        def to_float(v):
            """Parse numeric or range values and return a float.

            Accepts single numbers like "80" or ranges like "70-105" / "70 – 105".
            Uses the average for ranges. Normalizes Arabic/Persian digits and separators.
            """
            if v in (None, ''):
                return None
            s = str(v).strip()
            if s in {'-', '–', '—'}:
                return None
            try:
                # Normalize Arabic/Persian digits to ASCII
                digits_src = '٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹'
                digits_dst = '01234567890123456789'
                trans = str.maketrans({src: dst for src, dst in zip(digits_src, digits_dst)})
                s = s.translate(trans)

                # Normalize decimal separator and strip thousands separators
                s = s.replace('\u066b', '.').replace('٫', '.').replace('٬', '').replace(',', '')

                import re
                # استخرج الأرقام، واسمح بالسالب لدعم قيم سالبة إن وُجدت
                numbers = re.findall(r'-?\d+(?:\.\d+)?', s)
                if not numbers:
                    return None
                if len(numbers) == 1:
                    return float(numbers[0])
                a = float(numbers[0])
                b = float(numbers[1])
                return (a + b) / 2.0
            except Exception:
                return None

        price_housing = to_float(housing_val)
        price_commercial = to_float(commercial_val)
        price_industrial = to_float(industrial_val)
        price_agricultural = to_float(agricultural_val)
        legacy_price = to_float(legacy_price_val)

        if legacy_price is not None and not any(v is not None for v in (price_housing, price_commercial, price_industrial, price_agricultural)):
            price_housing = price_housing or legacy_price
            price_commercial = price_commercial or legacy_price
            price_industrial = price_industrial or legacy_price
            price_agricultural = price_agricultural or legacy_price

        if not wilaya_str or not region_str or all(v is None for v in (price_housing, price_commercial, price_industrial, price_agricultural, legacy_price)):
            skipped += 1
            continue

        existing = (
            CompanyLandPrice.query
            .filter_by(company_profile_id=profile.id, wilaya=wilaya_str, region=region_str)
            .first()
        )
        if existing:
            changed = False
            if price_housing is not None and existing.price_housing != price_housing:
                existing.price_housing = price_housing
                changed = True
            if price_commercial is not None and existing.price_commercial != price_commercial:
                existing.price_commercial = price_commercial
                changed = True
            if price_industrial is not None and existing.price_industrial != price_industrial:
                existing.price_industrial = price_industrial
                changed = True
            if price_agricultural is not None and existing.price_agricultural != price_agricultural:
                existing.price_agricultural = price_agricultural
                changed = True
            fallback_price = next((v for v in (price_housing, price_commercial, price_industrial, price_agricultural, legacy_price) if v is not None), None)
            if fallback_price is not None and (existing.price_per_sqm is None or existing.price_per_sqm != fallback_price):
                existing.price_per_sqm = fallback_price
                changed = True
            # حافظ على التوافق مع قواعد البيانات التي لديها عمود price_per_meter
            if has_price_per_meter and fallback_price is not None:
                current_val = getattr(existing, 'price_per_meter', None)
                if current_val is None or current_val != fallback_price:
                    existing.price_per_meter = fallback_price
                    changed = True
            if changed:
                updated += 1
        else:
            fallback_price = next((v for v in (price_housing, price_commercial, price_industrial, price_agricultural, legacy_price) if v is not None), None)
            new_obj = CompanyLandPrice(
                company_profile_id=profile.id,
                wilaya=wilaya_str,
                region=region_str,
                price_housing=price_housing,
                price_commercial=price_commercial,
                price_industrial=price_industrial,
                price_agricultural=price_agricultural,
                price_per_sqm=fallback_price,
            )
            if has_price_per_meter:
                new_obj.price_per_meter = fallback_price
            db.session.add(new_obj)
            inserted += 1

    db.session.commit()
    flash(f'تمت معالجة الملف: تمت إضافة {inserted} وتحديث {updated} وتجاوز {skipped} صف.', 'success')
    return redirect(url_for('company.edit_profile'))
