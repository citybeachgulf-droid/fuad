from flask import Blueprint, render_template, request, redirect, url_for, flash
from models import db, User, InviteToken, News, Advertisement, LandPrice, BankProfile
from flask_login import login_required, current_user
from urllib.parse import urljoin
from flask import current_app
from zoneinfo import ZoneInfo
import os
import time
from werkzeug.utils import secure_filename
from utils import store_file_and_get_url

admin_bp = Blueprint('admin', __name__, template_folder='../templates/admin', static_folder='../static')
# --- Logo upload helpers (admin) ---
ALLOWED_LOGO_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

def _allowed_logo_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_LOGO_EXTENSIONS



# --- Dashboard (محمي ويعرض البنوك والشركات) ---
from models import ValuationRequest

@admin_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403

    banks = User.query.filter_by(role='bank').all()
    companies = User.query.filter_by(role='company').all()
    clients = User.query.filter_by(role='client').all()
    requests = ValuationRequest.query.all()

    latest_news = News.query.order_by(News.created_at.desc()).limit(3).all()

    return render_template(
        'dashboard.html',
        banks=banks,
        companies=companies,
        clients=clients,
        requests=requests,
        latest_news=latest_news
    )

# --- صفحة عرض طلبات التثمين ---
@admin_bp.route('/requests')
@login_required
def requests_list():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    requests = ValuationRequest.query.order_by(ValuationRequest.id.desc()).all()
    return render_template('requests.html', requests=requests)

# --- إضافة بنك ---
@admin_bp.route('/add_bank', methods=['POST'])
@login_required
def add_bank():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    name = request.form['name']
    email = request.form['email']
    phone = request.form.get('phone')
    # إنشاء دعوة
    invite = InviteToken.generate(email=email, role='bank', invited_by_id=current_user.id, name=name, phone=phone)
    # إنشاء رابط التسجيل
    base_url = (f"http://{current_app.config['SERVER_NAME']}/" if current_app.config.get('SERVER_NAME') else request.host_url)
    invite_url = urljoin(base_url, url_for('auth.register', token=invite.token))

    flash(f'تم إنشاء دعوة للبنك. رابط التسجيل: {invite_url}', 'success')
    return redirect(url_for('admin.dashboard'))

# --- إضافة شركة تثمين ---
@admin_bp.route('/add_company', methods=['POST'])
@login_required
def add_company():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    name = request.form['name']
    email = request.form['email']
    phone = request.form.get('phone')
    # إنشاء دعوة
    invite = InviteToken.generate(email=email, role='company', invited_by_id=current_user.id, name=name, phone=phone)
    # إنشاء رابط التسجيل
    base_url = (f"http://{current_app.config['SERVER_NAME']}/" if current_app.config.get('SERVER_NAME') else request.host_url)
    invite_url = urljoin(base_url, url_for('auth.register', token=invite.token))

    flash(f'تم إنشاء دعوة للشركة. رابط التسجيل: {invite_url}', 'success')
    return redirect(url_for('admin.dashboard'))

# --- صفحة عرض البنوك ---
@admin_bp.route('/banks')
@login_required
def banks():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    banks = User.query.filter_by(role='bank').all()
    return render_template('banks.html', banks=banks)

# --- صفحة عرض شركات التثمين ---
@admin_bp.route('/companies')
@login_required
def companies():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    companies = User.query.filter_by(role='company').all()
    return render_template('companies.html', companies=companies)


# --- رفع أسعار الأراضي (إكسل) ---
@admin_bp.route('/land_prices/upload', methods=['POST'])
@login_required
def upload_land_prices():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403

    file = request.files.get('prices_file')
    if not file or not file.filename:
        flash('الرجاء اختيار ملف .xlsx أو .csv', 'danger')
        return redirect(url_for('admin.companies'))

    filename_lower = file.filename.lower()
    header_row = None
    rows_iter = None

    # Parse XLSX
    if filename_lower.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook  # lazy import
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                flash('ملف الإكسل فارغ.', 'danger')
                return redirect(url_for('admin.companies'))
        except ImportError:
            flash('مكتبة openpyxl غير متوفرة. يرجى رفع ملف CSV بدلاً من ذلك.', 'danger')
            return redirect(url_for('admin.companies'))
        except Exception:
            flash('تعذر قراءة ملف الإكسل. تأكد من أن الصيغة .xlsx صحيحة.', 'danger')
            return redirect(url_for('admin.companies'))

    # Parse CSV
    elif filename_lower.endswith('.csv'):
        import csv, io
        try:
            raw = file.stream.read()
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                try:
                    text = raw.decode('cp1256')
                except Exception:
                    text = raw.decode('latin1')
            reader = csv.reader(io.StringIO(text))
            header_row = next(reader, None)
            if header_row is None:
                flash('ملف CSV فارغ.', 'danger')
                return redirect(url_for('admin.companies'))
            rows_iter = reader
        except Exception:
            flash('تعذر قراءة ملف CSV.', 'danger')
            return redirect(url_for('admin.companies'))
    else:
        flash('صيغة غير مدعومة. الرجاء رفع .xlsx أو .csv', 'danger')
        return redirect(url_for('admin.companies'))

    def normalize_header_key(val):
        s = str(val or '').strip().lower()
        s = s.replace('ـ', '')
        import re
        # remove Arabic diacritics
        s = re.sub(r'[\u0617-\u061A\u064B-\u0652\u0670\u0653-\u065F]', '', s)
        # normalize common separators to spaces
        s = s.replace('_', ' ').replace('-', ' ').replace('/', ' ')
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    # Synonyms for supported headers (normalized)
    header_synonyms = {
        'wilaya': {
            'wilaya', 'الولاية', 'ولاية', 'ولايه', 'الولايه',
            'محافظة', 'المحافظة', 'governorate', 'state', 'province'
        },
        'region': {
            'region', 'المنطقة', 'منطقة', 'المنطقه', 'منطقه',
            'حي', 'الحي', 'حى', 'الحى', 'المدينة', 'مدينة',
            'district', 'area', 'neighborhood', 'neighbourhood', 'locality', 'city'
        },
        'housing': {'سكني', 'سكنية', 'سكن', 'housing'},
        'commercial': {'تجاري', 'تجارية', 'commercial'},
        'industrial': {'صناعي', 'صناعية', 'industrial'},
        'agricultural': {'زراعي', 'زراعية', 'agricultural', 'agriculture'},
        'price_per_sqm': {'price', 'price per sqm', 'price_per_sqm', 'سعر المتر', 'سعر_المتر', 'السعر'},
    }
    normalized_synonyms = {k: {normalize_header_key(v) for v in vs} for k, vs in header_synonyms.items()}

    header_map = {}
    for idx, col in enumerate(header_row):
        key = normalize_header_key(col)
        for canonical, syns in normalized_synonyms.items():
            if key in syns and canonical not in header_map:
                header_map[canonical] = idx
                break

    # مطلوب: الولاية والمنطقة، وأحد أعمدة (سكني/تجاري/صناعي/زراعي) أو السعر القديم
    if 'wilaya' not in header_map or 'region' not in header_map:
        flash('العناوين يجب أن تتضمن: الولاية، المنطقة', 'danger')
        return redirect(url_for('admin.companies'))
    if not any(k in header_map for k in ('housing', 'commercial', 'industrial', 'agricultural', 'price_per_sqm')):
        flash('العناوين يجب أن تتضمن أحد الأعمدة: سكني، تجاري، صناعي، زراعي (أو سعر المتر القديم)', 'danger')
        return redirect(url_for('admin.companies'))

    inserted = 0
    updated = 0
    skipped = 0

    # تحقق من وجود عمود price_per_meter في الجداول لملائمة الإصدارات القديمة
    try:
        from sqlalchemy import inspect as sa_inspect
        inspector = sa_inspect(db.engine)
        land_cols = [c['name'] for c in inspector.get_columns('land_prices')]
        has_price_per_meter = 'price_per_meter' in land_cols
    except Exception:
        has_price_per_meter = False

    for row in rows_iter:
        if row is None:
            skipped += 1
            continue

        wilaya_val = row[header_map['wilaya']] if len(row) > header_map['wilaya'] else None
        region_val = row[header_map['region']] if len(row) > header_map['region'] else None
        # استخراج أسعار حسب الاستعمال
        def get_val(col_key):
            if col_key in header_map and len(row) > header_map[col_key]:
                return row[header_map[col_key]]
            return None

        housing_val = get_val('housing')
        commercial_val = get_val('commercial')
        industrial_val = get_val('industrial')
        agricultural_val = get_val('agricultural')
        legacy_price_val = get_val('price_per_sqm')

        wilaya_str = str(wilaya_val or '').strip()
        region_str = str(region_val or '').strip()

        def to_float(v):
            """Parse a numeric cell that may contain a single value or a range.

            Supported examples:
            - "75"
            - "60-100", "60 – 100", "60 — 100", "60 الى 100", "60 إلى 100"
            - Arabic digits are normalized (٠١٢٣٤٥٦٧٨٩ / ۰۱۲۳۴۵۶۷۸۹)
            - Thousand separators like "," or "٬" are removed
            Returns the average when a range is provided.
            """
            if v in (None, ''):
                return None
            s = str(v).strip()
            # Treat placeholders for empty as None
            if s in {'-', '–', '—'}:
                return None
            try:
                # Normalize Arabic/Persian digits to ASCII
                digits_src = '٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹'
                digits_dst = '01234567890123456789'
                trans = str.maketrans({src: dst for src, dst in zip(digits_src, digits_dst)})
                s = s.translate(trans)

                # Normalize decimal separator and remove thousands separators
                s = s.replace('\u066b', '.').replace('٫', '.').replace('٬', '').replace(',', '')

                # Normalize common range connectors to a hyphen for readability
                # but use regex to extract numbers to be robust
                import re
                numbers = re.findall(r'-?\d+(?:\.\d+)?', s)
                if not numbers:
                    return None
                if len(numbers) == 1:
                    return float(numbers[0])
                # Average first two numbers if a range is given
                a = float(numbers[0])
                b = float(numbers[1])
                return (a + b) / 2.0
            except Exception:
                return None

        price_housing = to_float(housing_val)
        price_commercial = to_float(commercial_val)
        price_industrial = to_float(industrial_val)
        price_agricultural = to_float(agricultural_val)
        legacy_price = to_float(legacy_price_val)

        # إذا كان الملف قديماً بعمود واحد، عمّم السعر على كل الاستعمالات
        if legacy_price is not None and not any(v is not None for v in (price_housing, price_commercial, price_industrial, price_agricultural)):
            price_housing = price_housing or legacy_price
            price_commercial = price_commercial or legacy_price
            price_industrial = price_industrial or legacy_price
            price_agricultural = price_agricultural or legacy_price

        # يجب أن يتوفر على الأقل أحد الأسعار
        if not wilaya_str or not region_str or all(v is None for v in (price_housing, price_commercial, price_industrial, price_agricultural, legacy_price)):
            skipped += 1
            continue

        existing = LandPrice.query.filter_by(wilaya=wilaya_str, region=region_str).first()
        if existing:
            changed = False
            if price_housing is not None and existing.price_housing != price_housing:
                existing.price_housing = price_housing
                changed = True
            if price_commercial is not None and existing.price_commercial != price_commercial:
                existing.price_commercial = price_commercial
                changed = True
            if price_industrial is not None and existing.price_industrial != price_industrial:
                existing.price_industrial = price_industrial
                changed = True
            if price_agricultural is not None and existing.price_agricultural != price_agricultural:
                existing.price_agricultural = price_agricultural
                changed = True
            # حدّث السعر القديم إذا متاح ولم يكن موجوداً
            fallback_price = next((v for v in (price_housing, price_commercial, price_industrial, price_agricultural, legacy_price) if v is not None), None)
            if fallback_price is not None and (existing.price_per_sqm is None or existing.price_per_sqm != fallback_price):
                existing.price_per_sqm = fallback_price
                changed = True
            # حافظ على التوافق مع قواعد البيانات التي لديها عمود price_per_meter
            if has_price_per_meter and fallback_price is not None:
                current_val = getattr(existing, 'price_per_meter', None)
                if current_val is None or current_val != fallback_price:
                    existing.price_per_meter = fallback_price
                    changed = True
            if changed:
                updated += 1
        else:
            fallback_price = next((v for v in (price_housing, price_commercial, price_industrial, price_agricultural, legacy_price) if v is not None), None)
            new_obj = LandPrice(
                wilaya=wilaya_str,
                region=region_str,
                price_housing=price_housing,
                price_commercial=price_commercial,
                price_industrial=price_industrial,
                price_agricultural=price_agricultural,
                price_per_sqm=fallback_price,
            )
            if has_price_per_meter:
                new_obj.price_per_meter = fallback_price
            db.session.add(new_obj)
            inserted += 1

    db.session.commit()
    flash(f'تمت معالجة الملف: تمت إضافة {inserted} وتحديث {updated} وتجاوز {skipped} صف.', 'success')
    return redirect(url_for('admin.companies'))

# --- صفحة عرض العملاء ---
@admin_bp.route('/clients')
@login_required
def clients():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    clients = User.query.filter_by(role='client').all()
    return render_template('clients.html', clients=clients)

# --- صفحة عرض الدعوات ---
@admin_bp.route('/invites')
@login_required
def invites():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403

    invites_qs = InviteToken.query.order_by(InviteToken.created_at.desc()).all()

    base_url = (f"http://{current_app.config['SERVER_NAME']}/" if current_app.config.get('SERVER_NAME') else request.host_url)
    invites_data = []
    for inv in invites_qs:
        invite_url = urljoin(base_url, url_for('auth.register', token=inv.token))
        invites_data.append({
            'name': (inv.name or (inv.email.split('@')[0] if inv.email else '')),
            'email': inv.email,
            'role': inv.role,
            'url': invite_url,
            'used_at': inv.used_at,
            'expires_at': inv.expires_at,
        })

    return render_template('invites.html', invites=invites_data)


# --- إعداد رفع صور الأخبار ---
def _ensure_news_upload_dir(app):
    configured = app.config.get('NEWS_UPLOAD_FOLDER')
    if configured:
        news_upload = configured
    else:
        base_upload = app.config.get('UPLOAD_FOLDER', os.path.join(app.root_path, 'static', 'uploads'))
        news_upload = os.path.join(os.path.dirname(base_upload), 'news') if base_upload.endswith('logos') else os.path.join(base_upload, 'news')
    os.makedirs(news_upload, exist_ok=True)
    return news_upload


# --- قائمة الأخبار ---
@admin_bp.route('/news')
@login_required
def news_list():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    items = News.query.order_by(News.created_at.desc()).all()
    return render_template('news_list.html', news_list=items)


# --- إنشاء خبر جديد ---
@admin_bp.route('/news/new', methods=['GET', 'POST'])
@login_required
def news_new():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        body = request.form.get('body', '').strip()
        if not title:
            flash('الرجاء إدخال عنوان الخبر', 'danger')
            return redirect(url_for('admin.news_new'))

        image_path_rel = None
        file = request.files.get('image')
        if file and file.filename:
            filename = secure_filename(file.filename)
            upload_dir = _ensure_news_upload_dir(current_app)
            # ensure unique name if exists (for local fallback)
            base, ext = os.path.splitext(filename)
            counter = 1
            unique_filename = filename
            while os.path.exists(os.path.join(upload_dir, unique_filename)):
                unique_filename = f"{base}_{counter}{ext}"
                counter += 1
            object_key = f"uploads/news/{unique_filename}"
            stored = store_file_and_get_url(
                file,
                key=object_key,
                local_abs_dir=upload_dir,
                filename=unique_filename,
            )
            image_path_rel = stored

        news = News(title=title, body=body, image_path=image_path_rel)
        db.session.add(news)
        db.session.commit()
        flash('تم إضافة الخبر بنجاح', 'success')
        return redirect(url_for('admin.news_list'))

    return render_template('news_form.html')


# --- إدارة الإعلانات ---
@admin_bp.route('/ads')
@login_required
def ads_list():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    items = Advertisement.query.order_by(Advertisement.placement.asc(), Advertisement.sort_order.asc(), Advertisement.created_at.desc()).all()
    return render_template('ads_list.html', ads=items)


@admin_bp.route('/ads/new', methods=['GET', 'POST'])
@login_required
def ads_new():
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403

    if request.method == 'POST':
        title = (request.form.get('title') or '').strip()
        target_url = (request.form.get('target_url') or '').strip()
        placement = (request.form.get('placement') or 'homepage_top').strip()
        sort_order_raw = (request.form.get('sort_order') or '0').strip()
        is_active = (request.form.get('is_active') == 'on')
        start_at = request.form.get('start_at') or None
        end_at = request.form.get('end_at') or None

        # Parse datetimes if provided and normalize to UTC (naive)
        from datetime import datetime
        fmt = '%Y-%m-%dT%H:%M'
        tz_name = current_app.config.get('TIMEZONE', 'Asia/Muscat')
        local_tz = ZoneInfo(tz_name)
        utc_tz = ZoneInfo('UTC')

        def to_utc_naive(dt_str: str):
            if not dt_str:
                return None
            local_dt = datetime.strptime(dt_str, fmt).replace(tzinfo=local_tz)
            return local_dt.astimezone(utc_tz).replace(tzinfo=None)

        try:
            start_dt = to_utc_naive(start_at)
            end_dt = to_utc_naive(end_at)
            dates_in_utc = True
        except Exception:
            # Fallback to naive values without conversion
            start_dt = datetime.strptime(start_at, fmt) if start_at else None
            end_dt = datetime.strptime(end_at, fmt) if end_at else None
            dates_in_utc = False

        image_path_rel = None
        file = request.files.get('image')
        if file and file.filename:
            filename = secure_filename(file.filename)
            upload_dir = _ensure_ads_upload_dir(current_app)
            base, ext = os.path.splitext(filename)
            counter = 1
            unique_filename = filename
            while os.path.exists(os.path.join(upload_dir, unique_filename)):
                unique_filename = f"{base}_{counter}{ext}"
                counter += 1
            object_key = f"uploads/ads/{unique_filename}"
            stored = store_file_and_get_url(
                file,
                key=object_key,
                local_abs_dir=upload_dir,
                filename=unique_filename,
            )
            image_path_rel = stored

        try:
            sort_order = int(sort_order_raw)
        except Exception:
            sort_order = 0

        if not image_path_rel:
            flash('الرجاء رفع صورة للإعلان', 'danger')
            return redirect(url_for('admin.ads_new'))

        ad = Advertisement(
            title=title or None,
            image_path=image_path_rel,
            target_url=target_url or None,
            placement=placement or 'homepage_top',
            sort_order=sort_order,
            is_active=is_active,
            start_at=start_dt,
            end_at=end_dt,
            stored_in_utc=dates_in_utc,
        )
        db.session.add(ad)
        db.session.commit()
        flash('تم إنشاء الإعلان بنجاح', 'success')
        return redirect(url_for('admin.ads_list'))

    return render_template('ads_form.html')


@admin_bp.route('/ads/<int:ad_id>/toggle', methods=['POST'])
@login_required
def ads_toggle(ad_id: int):
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    ad = Advertisement.query.get_or_404(ad_id)
    ad.is_active = not ad.is_active
    db.session.commit()
    flash('تم تحديث حالة الإعلان', 'success')
    return redirect(url_for('admin.ads_list'))


@admin_bp.route('/ads/<int:ad_id>/delete', methods=['POST'])
@login_required
def ads_delete(ad_id: int):
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403
    ad = Advertisement.query.get_or_404(ad_id)
    db.session.delete(ad)
    db.session.commit()
    flash('تم حذف الإعلان', 'success')
    return redirect(url_for('admin.ads_list'))


def _ensure_ads_upload_dir(app):
    configured = app.config.get('ADS_UPLOAD_FOLDER')
    if configured:
        ads_upload = configured
    else:
        base_upload = app.config.get('UPLOAD_FOLDER', os.path.join(app.root_path, 'static', 'uploads'))
        ads_upload = os.path.join(os.path.dirname(base_upload), 'ads') if base_upload.endswith('logos') else os.path.join(base_upload, 'ads')
    os.makedirs(ads_upload, exist_ok=True)
    return ads_upload


# --- تحديث بيانات بنك (المدير فقط) ---
@admin_bp.route('/banks/<int:bank_id>/update', methods=['POST'])
@login_required
def update_bank(bank_id: int):
    if current_user.role != 'admin':
        return "غير مصرح لك بالوصول", 403

    bank_user = User.query.filter_by(id=bank_id, role='bank').first_or_404()

    name = (request.form.get('name') or '').strip()
    email = (request.form.get('email') or '').strip()
    phone = (request.form.get('phone') or '').strip()
    password = request.form.get('password') or ''

    # التحقق من البريد الإلكتروني الفريد إذا تم تغييره
    if email and email != bank_user.email:
        if User.query.filter_by(email=email).first():
            flash('البريد الإلكتروني مستخدم من قبل.', 'danger')
            return redirect(url_for('admin.banks'))

    if name:
        bank_user.name = name
    if email:
        bank_user.email = email
    if phone:
        bank_user.phone = phone
    else:
        # السماح بتفريغ الهاتف
        bank_user.phone = None

    if password:
        bank_user.set_password(password)

    # معالجة رفع الشعار من قبل المدير
    file = request.files.get('logo')
    if file and file.filename:
        if not _allowed_logo_file(file.filename):
            flash('صيغة الشعار غير مدعومة', 'danger')
            return redirect(url_for('admin.banks'))

        upload_dir = current_app.config.get('UPLOAD_FOLDER')
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"bank_{bank_user.id}_{int(time.time())}_" + secure_filename(file.filename)
        save_path = os.path.join(upload_dir, filename)
        file.save(save_path)

        # ضمان وجود ملف تعريف للبنك
        profile = bank_user.bank_profile
        if profile is None:
            profile = BankProfile(user_id=bank_user.id, slug=f"bank-{bank_user.id}")
            db.session.add(profile)
            db.session.flush()

        rel_path = os.path.relpath(save_path, os.path.join(current_app.root_path, 'static'))
        profile.logo_path = rel_path.replace('\\', '/')

    db.session.commit()
    flash('تم تحديث بيانات البنك بنجاح', 'success')
    return redirect(url_for('admin.banks'))
